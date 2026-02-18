# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import timedelta, date
import base64
import io
import json

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
except ImportError:
    openpyxl = None

from dateutil.relativedelta import relativedelta

class BuildingParametricReportWizard(models.TransientModel):
    _name = 'building.parametric.report.wizard'
    _description = 'Wizard Reporte Paramétrico (Flujo Financiero)'

    # === CONFIGURACIÓN ===
    work_id = fields.Many2one('building.work', string='Obra', required=True, readonly=True)
    budget_id = fields.Many2one(
        'building.budget', 
        string='Presupuesto', 
        required=True,
        domain="[('work_id', '=', work_id), ('state', '=', 'validated')]"
    )
    
    period_type = fields.Selection([
        ('month', 'Mensual'),
        ('biweekly', 'Quincenal'),
        ('week', 'Semanal')
    ], string='Periodo', default='month', required=True)

    # === COLUMNAS A INCLUIR ===
    col_budgeted = fields.Boolean(string='Presupuestado', default=True)
    col_advance = fields.Boolean(string='Anticipo', default=True, help='Muestra la columna de anticipo definido en cada partida')
    
    col_budgeted_accum = fields.Boolean(string='Acumulado Presupuestado', default=True)
    
    col_executed = fields.Boolean(string='Ejecutado Real', default=False)
    col_executed_accum = fields.Boolean(string='Acumulado Ejecutado', default=False)
    col_difference = fields.Boolean(string='Diferencia', default=False)

    # === RESULTADOS ===
    html_content = fields.Html(string='Vista Previa', readonly=True)
    excel_file = fields.Binary(string='Archivo Excel')
    excel_filename = fields.Char(string='Nombre Archivo')
    report_data_json = fields.Text(string='JSON Data', help='Almacena datos temporales')

    def _get_periods(self, start_date, end_date):
        """Genera lista de periodos entre start y end según period_type."""
        periods = []
        current = start_date
        
        while current <= end_date:
            if self.period_type == 'month':
                # Mensual: 1ro al último del mes
                p_start = current.replace(day=1)
                next_month = (p_start.replace(day=28) + timedelta(days=4)).replace(day=1)
                p_end = next_month - timedelta(days=1)
                label = p_start.strftime('%b %Y').title()
                current = next_month
            
            elif self.period_type == 'biweekly':
                # Quincenal: 1-15 y 16-fin
                if current.day <= 15:
                    p_start = current.replace(day=1)
                    p_end = current.replace(day=15)
                    label = f"1Q {p_start.strftime('%b %y').title()}"
                    current = current.replace(day=16)
                else:
                    p_start = current.replace(day=16)
                    next_month = (p_start.replace(day=28) + timedelta(days=4)).replace(day=1)
                    p_end = next_month - timedelta(days=1)
                    label = f"2Q {p_start.strftime('%b %y').title()}"
                    current = next_month

            elif self.period_type == 'week':
                # Semanal: Lunes a Domingo
                p_start = current - timedelta(days=current.weekday())
                p_end = p_start + timedelta(days=6)
                label = f"Sem {p_start.isocalendar()[1]}"
                current = p_end + timedelta(days=1)

            periods.append({
                'start': p_start, 
                'end': p_end, 
                'label': label,
                'key': len(periods) # index 0, 1, 2...
            })
            
            if len(periods) > 100: # Safety break
                break
                
        return periods





    def _get_report_data(self):
        """Calcula toda la data del reporte."""
        if not self.budget_id:
            return {}

        # 0. Determinar Fecha Inicio de Obra
        # Prioridad: Fecha Inicio Contrato > Fecha Creación Obra > Hoy
        start_date = self.work_id.contract_date_start or self.work_id.create_date.date() or date.today()
        # Ajustar al día 1 del mes para alineación mensual
        base_date = start_date.replace(day=1)

        # 1. Rango de fechas global basado en periodos (Enteros)
        # Buscar min y max periodos en las líneas
        lines = self.budget_id.chapter_ids.line_ids
        if not lines:
            min_period = 1
            max_period = self.budget_id.duration_months or 12
        else:
            # Filtrar 0 o False si los hubiera
            periods = lines.mapped('period_from') + lines.mapped('period_to')
            periods = [p for p in periods if p]
            if periods:
                min_period = min(periods)
                max_period = max(periods)
            else:
                min_period = 1
                max_period = 12

        # Convertir periodos enteros a Fechas
        # Periodo 1 = Mes base_date
        # Periodo N = base_date + (N-1) meses
        min_date = base_date + relativedelta(months=min_period - 1)
        max_date = base_date + relativedelta(months=max_period) - timedelta(days=1)

        # 2. Generar Periodos Columnas (Wizard Period Type)
        periods_cols = self._get_periods(min_date, max_date)

        # 3. Estructura de Capítulos y Partidas
        chapters_data = []
        grand_totals = {
            'amount': 0.0,
            'advance': 0.0,
            'per_period': {p['key']: 0.0 for p in periods_cols},
            'accumulated': {p['key']: 0.0 for p in periods_cols},
            'executed_per_period': {p['key']: 0.0 for p in periods_cols},
            'executed_accum': {p['key']: 0.0 for p in periods_cols},
            'diff_per_period': {p['key']: 0.0 for p in periods_cols},
        }

        accum_total = 0.0
        accum_executed = 0.0

        for chapter in self.budget_id.chapter_ids:
            chap_dict = {
                'code': chapter.code,
                'name': chapter.name,
                'amount': 0.0,
                'advance': 0.0,
                'lines': [],
                'per_period': {p['key']: 0.0 for p in periods_cols},
                'executed_per_period': {p['key']: 0.0 for p in periods_cols},
            }

            for line in chapter.line_ids:
                line_amount = line.amount
                line_advance = line.advance if self.col_advance else 0.0
                
                # Distribución Presupuestada
                dist_amounts = {}
                
                # Convertir periodos de la línea a fechas reales
                l_p_from = line.period_from or 1
                l_p_to = line.period_to or 1
                
                line_date_from = base_date + relativedelta(months=l_p_from - 1)
                # Fin de mes del periodo final
                line_date_to = base_date + relativedelta(months=l_p_to) - timedelta(days=1)
                
                total_days = (line_date_to - line_date_from).days + 1
                daily_rate = line_amount / total_days if total_days > 0 else 0

                for p in periods_cols:
                    overlap_start = max(p['start'], line_date_from)
                    overlap_end = min(p['end'], line_date_to)
                    
                    if overlap_start <= overlap_end:
                        days = (overlap_end - overlap_start).days + 1
                        amt = daily_rate * days
                    else:
                        amt = 0.0
                        
                    dist_amounts[p['key']] = amt
                    chap_dict['per_period'][p['key']] += amt
                    grand_totals['per_period'][p['key']] += amt

                # Ejecutado Real
                exec_amounts = {}
                if self.col_executed:
                    for p in periods_cols:
                        # TODO: Lógica de ejecutado global 
                        exec_amounts[p['key']] = 0.0 
                        
                line_dict = {
                    'code': line.code,
                    'name': line.name,
                    'amount': line_amount,
                    'advance': line_advance,
                    'dist': dist_amounts,
                    'exec': exec_amounts
                }
                chap_dict['lines'].append(line_dict)
                chap_dict['amount'] += line_amount
                chap_dict['advance'] += line_advance

            chapters_data.append(chap_dict)
            grand_totals['amount'] += chap_dict['amount']
            grand_totals['advance'] += chap_dict['advance']

        # Calcular Acumulados Globales
        run_accum = 0.0
        run_exec = 0.0
        for p in periods_cols:
            run_accum += grand_totals['per_period'][p['key']]
            grand_totals['accumulated'][p['key']] = run_accum
            
            if self.col_executed:
                real_global = self.env['building.real.line'].search([
                    ('work_id', '=', self.work_id.id),
                    ('date', '>=', p['start']),
                    ('date', '<=', p['end'])
                ])
                amt_exec = sum(real_global.mapped('amount'))
                grand_totals['executed_per_period'][p['key']] = amt_exec
                run_exec += amt_exec
                grand_totals['executed_accum'][p['key']] = run_exec
                grand_totals['diff_per_period'][p['key']] = grand_totals['per_period'][p['key']] - amt_exec

        return {
            'work_name': self.work_id.name,
            'budget_name': self.budget_id.name,
            'periods': periods_cols,
            'chapters': chapters_data,
            'totals': grand_totals
        }

    def action_generate_report(self):
        """Genera el reporte HTML."""
        data = self._get_report_data()
        self.report_data_json = json.dumps(data, default=str)
        self.html_content = self._generate_html_table(data)
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'building.parametric.report.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
            'context': self.env.context,
        }

    def _generate_html_table(self, data):
        html = """
        <style>
            .parametric-table { border-collapse: collapse; width: 100%; font-size: 12px; font-family: Roboto, sans-serif; }
            .parametric-table th { background-color: #2c3e50; color: white; padding: 8px; text-align: center; border: 1px solid #ddd; position: sticky; top: 0; z-index: 10; }
            .parametric-table td { padding: 5px; border: 1px solid #ddd; text-align: right; }
            .row-chapter { background-color: #ecf0f1; font-weight: bold; text-align: left; }
            .row-chapter td { font-weight: bold !important; border-top: 2px solid #bdc3c7; }
            .row-total { background-color: #fcf3cf; font-weight: bold; }
            .col-text { text-align: left; }
            .num { white-space: nowrap; }
        </style>
        <div style="overflow-x: auto; max-height: 70vh;">
            <table class="parametric-table">
                <thead>
                    <tr>
                        <th style="min-width:80px">CL.</th>
                        <th style="min-width:250px">CONCEPTO</th>
                        <th style="min-width:100px">IMPORTE</th>
                        """ 
        if self.col_advance:
            html += '<th style="min-width:100px">ANTICIPO</th>'
            
        for p in data['periods']:
            html += f'<th style="min-width:100px">{p["label"]}</th>'

        html += """</tr>
                </thead>
                <tbody>"""

        # Capítulos
        for chap in data['chapters']:
            html += f"""
                <tr class="row-chapter">
                    <td class="col-text"><b>{chap['code']}</b></td>
                    <td class="col-text"><b>{chap['name']}</b></td>
                    <td class="num"><b>{self._fmt(chap['amount'])}</b></td>
                    """
            if self.col_advance:
                html += f'<td class="num">{self._fmt(chap["advance"])}</td>'
            
            for p in data['periods']:
                val = chap['per_period'].get(p['key'], 0.0)
                html += f'<td class="num">{self._fmt(val)}</td>'
            
            html += "</tr>"

            # Partidas
            for line in chap['lines']:
                html += f"""
                    <tr>
                        <td class="col-text">{line['code']}</td>
                        <td class="col-text">{line['name']}</td>
                        <td class="num">{self._fmt(line['amount'])}</td>
                        """
                if self.col_advance:
                    html += f'<td class="num">{self._fmt(line["advance"])}</td>'
                
                for p in data['periods']:
                    val = line['dist'].get(p['key'], 0.0)
                    html += f'<td class="num">{self._fmt(val)}</td>'
                html += "</tr>"

        # Totales
        html += '<tr class="row-total"><td colspan="2">TOTAL</td>'
        html += f'<td class="num">{self._fmt(data["totals"]["amount"])}</td>'
        if self.col_advance:
            html += f'<td class="num">{self._fmt(data["totals"]["advance"])}</td>'
        
        for p in data['periods']:
            html += f'<td class="num">{self._fmt(data["totals"]["per_period"][p["key"]])}</td>'
        html += "</tr>"

        # Acumulado
        if self.col_budgeted_accum:
            html += '<tr style="background-color: #d5f5e3; font-weight: bold;"><td colspan="2">ACUMULADO</td><td></td>'
            if self.col_advance: html += '<td></td>'
            for p in data['periods']:
                html += f'<td class="num">{self._fmt(data["totals"]["accumulated"][p["key"]])}</td>'
            html += "</tr>"

        html += "</tbody></table></div>"
        return html

    def _fmt(self, amount):
        return "{:,.2f}".format(amount)

    def action_download_excel(self):
        """Genera y descarga el Excel."""
        if not openpyxl:
            raise UserError(_('La librería openpyxl no está instalada.'))

        data = self._get_report_data()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Flujo Financiero"
        ws.sheet_view.showGridLines = False  # Hide gridlines for cleaner look

        # --- Styles Definition ---
        # Colors
        COLOR_PRIMARY = "1F4E78"    # Dark Blue (Titles/Headers)
        COLOR_SECONDARY = "DDEBF7"  # Light Blue (Chapters)
        COLOR_TOTAL = "FFF2CC"      # Light Yellow (Totals)
        COLOR_ZEBRA = "F2F2F2"      # Very Light Grey
        COLOR_TEXT = "000000"
        COLOR_WHITE = "FFFFFF"

        # Fonts
        font_title = Font(name='Calibri', size=16, bold=True, color=COLOR_PRIMARY)
        font_subtitle = Font(name='Calibri', size=12, bold=True, color=COLOR_TEXT)
        font_header = Font(name='Calibri', size=11, bold=True, color=COLOR_WHITE)
        font_chapter = Font(name='Calibri', size=11, bold=True, color=COLOR_PRIMARY) # Dark text for chapters
        font_normal = Font(name='Calibri', size=11, color=COLOR_TEXT)
        font_total = Font(name='Calibri', size=11, bold=True, color=COLOR_TEXT)

        # Fills
        fill_header = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
        fill_chapter = PatternFill(start_color=COLOR_SECONDARY, end_color=COLOR_SECONDARY, fill_type="solid")
        fill_total = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")
        fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
        fill_none = PatternFill(fill_type=None)

        # Borders
        border_thin = Side(border_style="thin", color="AAAAAA")
        border_medium = Side(border_style="medium", color=COLOR_PRIMARY)
        
        border_frame = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
        border_bottom_thick = Border(bottom=border_medium)

        # Alignment
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')

        # Number Formats
        fmt_currency = '"$"#,##0.00_-'
        
        # --- Content Generation ---

        # 1. Title Section
        ws.merge_cells('B2:E2')
        ws['B2'] = "CONTROL DE OBRAS - FLUJO FINANCIERO"
        ws['B2'].font = font_title
        ws['B2'].alignment = align_left

        ws['B4'] = "OBRA:"
        ws['C4'] = data['work_name']
        ws['B4'].font = font_subtitle
        ws['C4'].font = font_normal

        ws['B5'] = "PRESUPUESTO:"
        ws['C5'] = data['budget_name']
        ws['B5'].font = font_subtitle
        ws['C5'].font = font_normal

        ws['B6'] = "FECHA:"
        ws['C6'] = date.today().strftime('%d/%m/%Y')
        ws['B6'].font = font_subtitle
        ws['C6'].font = font_normal

        # 2. Table Headers (Row 8)
        start_row = 8
        headers = ["CL.", "CONCEPTO", "IMPORTE"]
        if self.col_advance: headers.append("ANTICIPO")
        for p in data['periods']: headers.append(p['label'])

        max_col = len(headers) + 1 # +1 because we start at col B (index 2)
        
        for idx, title in enumerate(headers):
            col_idx = idx + 2 # Start at Column B
            cell = ws.cell(row=start_row, column=col_idx, value=title)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_frame
            
            # Widths
            col_letter = get_column_letter(col_idx)
            if idx == 0: ws.column_dimensions[col_letter].width = 12
            elif idx == 1: ws.column_dimensions[col_letter].width = 45 # Concepto wider
            else: ws.column_dimensions[col_letter].width = 18

        ws.freeze_panes = ws.cell(row=start_row+1, column=4) 

        # --- Data Rows ---
        row = start_row + 1
        
        for chap in data['chapters']:
            # Chapter Row
            # Code
            c = ws.cell(row=row, column=2, value=chap['code'])
            c.font = font_chapter
            c.fill = fill_chapter
            c.border = border_frame
            
            # Name
            c = ws.cell(row=row, column=3, value=chap['name'])
            c.font = font_chapter
            c.fill = fill_chapter
            c.border = border_frame

            # Amount
            c = ws.cell(row=row, column=4, value=chap['amount'])
            c.font = font_chapter
            c.fill = fill_chapter
            c.border = border_frame
            c.number_format = fmt_currency
            
            col_off = 5
            if self.col_advance:
                c = ws.cell(row=row, column=col_off, value=chap['advance'])
                c.font = font_chapter
                c.fill = fill_chapter
                c.border = border_frame
                c.number_format = fmt_currency
                col_off += 1

            for p in data['periods']:
                val = chap['per_period'].get(p['key'], 0.0)
                c = ws.cell(row=row, column=col_off, value=val)
                c.font = font_chapter
                c.fill = fill_chapter
                c.border = border_frame
                c.number_format = fmt_currency
                col_off += 1
            
            row += 1

            # Lines (Grouped under Chapter)
            line_idx = 0
            for line in chap['lines']:
                ws.row_dimensions[row].outlineLevel = 1 # Grouping
                ws.row_dimensions[row].hidden = False

                # Zebra Striping
                current_fill = fill_zebra if line_idx % 2 != 0 else fill_none

                c = ws.cell(row=row, column=2, value=line['code'])
                c.font = font_normal
                c.fill = current_fill
                c.border = border_frame
                
                c = ws.cell(row=row, column=3, value=line['name'])
                c.font = font_normal
                c.fill = current_fill
                c.border = border_frame
                
                c = ws.cell(row=row, column=4, value=line['amount'])
                c.font = font_normal
                c.fill = current_fill
                c.border = border_frame
                c.number_format = fmt_currency
                
                col_off = 5
                if self.col_advance:
                    c = ws.cell(row=row, column=col_off, value=line['advance'])
                    c.font = font_normal
                    c.fill = current_fill
                    c.border = border_frame
                    c.number_format = fmt_currency
                    col_off += 1
                
                for p in data['periods']:
                    val = line['dist'].get(p['key'], 0.0)
                    c = ws.cell(row=row, column=col_off, value=val)
                    c.font = font_normal
                    c.fill = current_fill
                    c.border = border_frame
                    c.number_format = fmt_currency
                    col_off += 1
                
                row += 1
                line_idx += 1

        # --- Totals Row ---
        c = ws.cell(row=row, column=2, value="TOTAL GENERAL")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        c.font = font_total
        c.fill = fill_total
        c.border = Border(top=border_medium, bottom=border_medium, left=border_thin, right=border_thin)
        c.alignment = align_right
        
        # Applying style to the merged cell's partner to ensure border shows
        c2 = ws.cell(row=row, column=3)
        c2.border = Border(top=border_medium, bottom=border_medium, right=border_thin)
        c2.fill = fill_total

        c = ws.cell(row=row, column=4, value=data["totals"]["amount"])
        c.font = font_total
        c.fill = fill_total
        c.border = Border(top=border_medium, bottom=border_medium, left=border_thin, right=border_thin)
        c.number_format = fmt_currency
        
        col_off = 5
        if self.col_advance:
            c = ws.cell(row=row, column=col_off, value=data["totals"]["advance"])
            c.font = font_total
            c.fill = fill_total
            c.border = Border(top=border_medium, bottom=border_medium, left=border_thin, right=border_thin)
            c.number_format = fmt_currency
            col_off += 1
        
        for p in data['periods']:
            val = data["totals"]["per_period"][p["key"]]
            c = ws.cell(row=row, column=col_off, value=val)
            c.font = font_total
            c.fill = fill_total
            c.border = Border(top=border_medium, bottom=border_medium, left=border_thin, right=border_thin)
            c.number_format = fmt_currency
            col_off += 1

        # Generar Archivo
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        self.excel_file = base64.b64encode(fp.read())
        self.excel_filename = f"Flujo_Financiero_{data['work_name']}.xlsx"
        fp.close()

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content?model=building.parametric.report.wizard&id={self.id}&field=excel_file&filename_field=excel_filename&download=true',
            'target': 'self',
        }

def custom_strftime(date_obj, fmt):
    # Simple strftime wrapper for Spanish names if needed, 
    # relying on Odoo/Python locale usually, but hardcoding for robustness here
    months = {
        1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
        7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
    }
    if '%b' in fmt:
        return fmt.replace('%b', months[date_obj.month]).replace('%Y', str(date_obj.year)).replace('%y', str(date_obj.year)[-2:])
    return date_obj.strftime(fmt)
